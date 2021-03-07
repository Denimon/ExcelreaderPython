import tkinter as tk
from openpyxl import load_workbook


workbook = load_workbook('/Users/denimon/Desktop/Stencil_program/2020.xlsx')

def only_numbers(char): 
    return char.replace(".", "0", 1).isdigit()


#Mainwindow
mainWindow = tk.Tk()
mainWindow.title("Stencilplats")
mainWindow.geometry("750x500")


#Header label
infoLabel = tk.Label(mainWindow, width=22, text="Hitta stencilplats", anchor='w')
infoLabel.config(font=("Raleway",18))
infoLabel.pack(side=tk.TOP, pady = 30)

row = tk.Frame(mainWindow)
articleLabel = tk.Label(row, width=22, text="Artikelnummer:", anchor='w')
articleLabel.config(font=("Raleway",16))
 
validation = row.register(only_numbers)
articleEntry = tk.Entry(row, validate="key", validatecommand=(validation, '%S'))
articleEntry.config(font=("Raleway",16))
 
row.pack(side=tk.TOP, fill=tk.X, padx=5, pady=5)
articleLabel.pack(side=tk.LEFT)
articleEntry.pack(side=tk.RIGHT, expand=tk.YES, fill=tk.X)

row2 = tk.Frame(mainWindow)
label3 = tk.Label(row2, width=60, text = "", anchor = 'w')
label3.config(font=("Raleway",18))
label3.pack(side=tk.LEFT, padx=150, ipady = 100)
#label2 = tk.Label(row2, width=22, text="Number 2", anchor='w')
#label2.config(font=("Raleway",16))
#validation = row2.register(only_numbers)
#number2 = tk.Entry(row2, validate="key", validatecommand=(validation, '%S'))
#number2.config(font=("Raleway",16))
 
row2.pack(side=tk.TOP, fill=tk.X, padx=5, pady=5)
label3.pack(side=tk.LEFT)
#number2.pack(side=tk.RIGHT, expand=tk.YES, fill=tk.X)


b1 = tk.Button(mainWindow, width = 15, text='Sök', background = "Light green", command=(lambda: searchStencil()))
b1.config(font=("Raleway", 18))   
b1.pack(side=tk.BOTTOM, padx=5, pady=50)

def searchStencil():

    searchInput = int(articleEntry.get())

    for row in workbook.worksheets[0].iter_rows():
        for cell in row:
            if cell.value == searchInput:
                label3.config(text=row[0].value)
            
                if cell.comment:
                    print(cell.comment.text)

                return    

    
    label3.config(text = "Hittade ingen stencilplats för {}. Försök igen".format(searchInput))
    return 

#def ():
 #   added_val = int(number1.get()) +  int(number2.get())
  #  label3 = tk.Label(root, width=60, text =  "SUM " + str(added_val), anchor = 'w')
   # label3.config(font=("Raleway",18))
    #label3.pack(side=tk.LEFT, padx=150, ipady = 100)

mainWindow.mainloop()